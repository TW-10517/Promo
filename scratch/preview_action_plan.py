import sys
from pathlib import Path
import openpyxl

xlsx_path = Path("projects/GALLERIA_Autumn_2026_Promotion_20260814_172306_cb7bf3/output/demo_planning_document_action_plan_20260814-172512.xlsx")
if not xlsx_path.is_file():
    print("File not found")
    sys.exit(0)

wb = openpyxl.load_workbook(str(xlsx_path))
ws = wb.active

out = ["=== GENERATED ACTION PLAN PREVIEW ==="]
out.append(f"Sheet Name: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}\n")

# Print first 12 rows
for row in ws.iter_rows(min_row=1, max_row=min(12, ws.max_row), values_only=True):
    row_str = " | ".join(str(c) if c is not None else "" for c in row)
    out.append(row_str)

sys.stdout.buffer.write(("\n".join(out) + "\n").encode("utf-8"))
