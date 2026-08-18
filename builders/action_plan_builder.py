"""
Action Plan builder: JSON -> .xlsx

The grid is fixed by ``templates/action_plan.json``:

* one column per template column (the first one is the fixed row label)
* one row per template row, in template order

Most cells are expected to stay blank -- they are filled in by hand after the
Promotion Department and the Planning Department have agreed the details. This
builder therefore never substitutes a placeholder for an empty value; it just
leaves the cell empty and highlights it lightly so reviewers can see at a
glance what still needs input.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.missing_tracker import MissingItem

HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")   # light blue
LABEL_FILL = PatternFill("solid", fgColor="F2F2F2")    # light grey
PENDING_FILL = PatternFill("solid", fgColor="FFF7E6")  # pale amber: to be filled in later
THIN = Side(style="thin", color="A6A6A6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BODY_FONT = "Yu Gothic"


def detect_missing(data: dict, template: dict) -> list[MissingItem]:
    """
    Find every non-fixed cell that would be written blank.

    Mirrors the fill logic in ``build()`` exactly (same row/column pairing,
    same "blank is expected, just flag it" rule) so the in-memory preview
    pipeline in ``core.router`` reports the same missing cells the exported
    ``.xlsx`` would highlight.
    """
    missing: list[MissingItem] = []
    columns = template.get("columns", [])
    template_rows = template.get("rows", [])
    ai_rows = data.get("rows", [])
    if not isinstance(ai_rows, list):
        ai_rows = []

    for offset, template_row in enumerate(template_rows):
        ai_row = ai_rows[offset] if offset < len(ai_rows) else {}
        if not isinstance(ai_row, dict):
            ai_row = {}
        row_label = str(template_row.get("item", "") or f"行{offset + 1}")

        for column in columns:
            if column.get("fixed"):
                continue
            value = str(ai_row.get(column["key"], "") or "").strip()
            if not value:
                missing.append(MissingItem(field=column.get("header", column["key"]), context=row_label))

    return missing


def build(data: dict, template: dict, output_path: str | Path) -> tuple[Path, list[MissingItem]]:
    """
    Write the Action Plan workbook.

    Parameters
    ----------
    data:
        Parsed JSON response: ``{"rows": [{<column key>: value, ...}, ...]}``.
    template:
        The loaded ``action_plan.json`` structure.
    output_path:
        Full destination path (including the .xlsx extension).

    Returns
    -------
    tuple[Path, list[MissingItem]]
        The written file, and every non-fixed cell left blank -- expected,
        since most of this grid is filled in by hand after the Promotion and
        Planning Departments confer, but tracked so nothing gets forgotten.
        Generation is never blocked by this list; it is purely informational.
    """
    path = Path(output_path)
    missing = detect_missing(data, template)
    columns = template.get("columns", [])
    template_rows = template.get("rows", [])
    ai_rows = data.get("rows", [])
    if not isinstance(ai_rows, list):
        ai_rows = []

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = template.get("sheet_name", "実施計画")

    # --- Document title (merged across all columns) ------------------------
    title = template.get("document_title", "実施計画書")
    sheet.cell(row=1, column=1, value=title)
    sheet.cell(row=1, column=1).font = Font(name=BODY_FONT, bold=True, size=14)
    if len(columns) > 1:
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(columns)
        )
    sheet.row_dimensions[1].height = 24

    # --- Header row --------------------------------------------------------
    header_row = 3
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=column.get("header", ""))
        cell.font = Font(name=BODY_FONT, bold=True)
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = column.get("width", 20)
    sheet.row_dimensions[header_row].height = 20

    # --- Data rows ---------------------------------------------------------
    # Pair by index: the AI is instructed to return exactly one entry per
    # template row, in order. If it returns fewer, the remaining rows are
    # written blank rather than shifted -- the fixed structure always wins.
    for offset, template_row in enumerate(template_rows):
        excel_row = header_row + 1 + offset
        ai_row = ai_rows[offset] if offset < len(ai_rows) else {}
        if not isinstance(ai_row, dict):
            ai_row = {}
        row_label = str(template_row.get("item", "") or f"行{offset + 1}")

        for index, column in enumerate(columns, start=1):
            key = column["key"]
            if column.get("fixed"):
                # Fixed label column: always taken from the template.
                value = template_row.get(key, "")
            else:
                value = str(ai_row.get(key, "") or "").strip()

            cell = sheet.cell(row=excel_row, column=index, value=value or None)
            cell.font = Font(name=BODY_FONT, bold=bool(column.get("fixed")))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if column.get("fixed"):
                cell.fill = LABEL_FILL
            elif not value:
                # Blank is the correct, expected state -- flag it as "to fill in"
                # visually (the `missing` list itself came from detect_missing()).
                cell.fill = PENDING_FILL

        sheet.row_dimensions[excel_row].height = 30

    # --- Legend ------------------------------------------------------------
    legend_row = header_row + len(template_rows) + 2
    legend = sheet.cell(
        row=legend_row,
        column=1,
        value="※ 色付きの空欄は、企画部との協議後に手入力で記入する欄です。",
    )
    legend.font = Font(name=BODY_FONT, size=9, italic=True)

    # Freeze the title and header so long plans stay readable.
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=2)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))
    return path, missing
